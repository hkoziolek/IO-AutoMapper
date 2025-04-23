import os
import pandas as pd
from openpyxl.styles import Font, PatternFill

def extract_lines_from_excel(file_path, columns, start_line, end_line):
    """Extract specified lines from an Excel file."""
    print(f"Reading from Excel file, row {start_line}-{end_line}...")
    try:
        data = pd.read_excel(file_path)
        # Select only the columns of interest
        data = data[columns]

        # Extract the specified range of rows (adjust for zero-based indexing)
        extracted_rows = data.iloc[start_line-1:end_line]
        if extracted_rows is None:
            print("Error extracting rows from Excel file.")
            return None, None
        return columns, extracted_rows
    except FileNotFoundError:
        print(f"Error extracting lines from Excel: The file '{file_path}' does not exist.")
        return None, None
    except Exception as e:
        print(f"Error extracting lines from Excel: {e}")
        return None, None

def check_ground_truth_mismatches(extracted_rows, out_df):
    """Check for mismatches between ground truth and generated data."""
    num_object_type_mismatch = 0
    num_process_object_type_mismatch = 0
    num_process_object_parameter_mismatch = 0
    
    for index, row in extracted_rows.iterrows():
        line = index + 2
        matching_row = out_df[out_df["Line"]==line]
        if matching_row.empty:
            print(f"Ground truth mismatch: Line {line} not found in output.")
        else:
            if row["Signal Tag \n(IO Tag)"] != matching_row["Signal Tag \n(IO Tag)"].values[0]:
                print(f"Ground truth mismatch: Signal Tag {row['Signal Tag \n(IO Tag)']} in line {line} does not match output.")

            # Check if matching_row["Signal Data Type"] is empty
            if matching_row["Signal Data Type"].values[0] != "" and row["CB Object Type"] != matching_row["Signal Data Type"].values[0]:
                print(f"Ground truth mismatch: CB Object Type {row['CB Object Type']} in line {line} does not match output {matching_row['Signal Data Type'].values[0]}.")
                num_object_type_mismatch += 1
                
            if matching_row["Function Block Type"].values[0] != "" and row["CB Process Object Type"] != matching_row["Function Block Type"].values[0]:
                print(f"Ground truth mismatch: CB Process Object Type {row['CB Process Object Type']} in line {line} does not match output {matching_row['Function Block Type'].values[0]}.")
                num_process_object_type_mismatch += 1
                
            if matching_row["Function Block Parameter"].values[0] != "" and row["CB Process Object Parameter"] != matching_row["Function Block Parameter"].values[0]:
                print(f"Ground truth mismatch: CB Process Object Parameter {row['CB Process Object Parameter']} in line {line} does not match output {matching_row['Function Block Parameter'].values[0]}.")
                num_process_object_parameter_mismatch += 1

    num_rows = extracted_rows.shape[0] # get number of rows in extracted_rows
    object_type_match_percentage = round((num_rows-num_object_type_mismatch) / num_rows * 100,2)
    process_object_type_match_percentage = round((num_rows-num_process_object_type_mismatch) / num_rows * 100,2)
    process_object_parameter_match_percentage = round((num_rows-num_process_object_parameter_mismatch) / num_rows * 100,2)

    print(f"Number of successful mappings:")
    print(f"Signal Data Type: {num_rows-num_object_type_mismatch}/{num_rows} ({object_type_match_percentage}%)")
    print(f"Function Block Type: {num_rows-num_process_object_type_mismatch}/{num_rows} ({process_object_type_match_percentage}%)")
    print(f"Function Block Parameter: {num_rows-num_process_object_parameter_mismatch}/{num_rows} ({process_object_parameter_match_percentage}%)")
    print("Ground truth check done.")

def check_ground_truth_mismatches2(extracted_rows, out_df, result_filename, signal_column_title):
    """Check for mismatches between ground truth and generated data."""
    num_object_type_mismatch = 0
    num_process_object_type_mismatch = 0
    num_process_object_parameter_mismatch = 0
    
    # Create a list to store all mismatch messages
    mismatch_messages = []
    
    for index, row in extracted_rows.iterrows():
        #line = index + 2
        line = row["Line"]
        matching_row = out_df[out_df["Line"]==line]
        if matching_row.empty:
            pass
            # message = f"Ground truth mismatch: Line {line} not found in output."
            # print(message)
            # mismatch_messages.append(message)
        else:
            if row[signal_column_title] != matching_row[signal_column_title].values[0]:
                message = f"Ground truth mismatch: Signal Tag {row['Signal Tag \n(IO Tag)']} in line {line} does not match output."
                print(message)
                mismatch_messages.append(message)

            # Check if matching_row["Signal Data Type"] is empty
            if matching_row["Signal Data Type"].values[0] != "" and row["Signal Data Type"] != matching_row["Signal Data Type"].values[0]:
                message = f"Ground truth mismatch: Signal Data Type {row['Signal Data Type']} in line {line} does not match output {matching_row['Signal Data Type'].values[0]}."
                print(message)
                mismatch_messages.append(message)
                num_object_type_mismatch += 1
                
            if matching_row["Function Block Type"].values[0] != "" and row["Function Block Type"] != matching_row["Function Block Type"].values[0]:
                message = f"Ground truth mismatch: Function Block Type {row['Function Block Type']} in line {line} does not match output {matching_row['Function Block Type'].values[0]}."
                print(message)
                mismatch_messages.append(message)
                num_process_object_type_mismatch += 1
                
            if matching_row["Function Block Parameter"].values[0] != "":
                # Convert pandas NaN to string "n/a" for comparison
                row_value = row["Function Block Parameter"]
                # Check if value is NaN and handle appropriately
                if pd.isna(row_value):
                    row_value = "n/a"  # Or whatever string representation you want for NaN values
                
                if str(row_value).strip() != str(matching_row["Function Block Parameter"].values[0]).strip():
                    message = f"Ground truth mismatch: Function Block Parameter '{row_value}' in line {line} does not match output '{matching_row['Function Block Parameter'].values[0]}'."
                    print(message)
                    mismatch_messages.append(message)
                    num_process_object_parameter_mismatch += 1
    
    num_rows = extracted_rows.shape[0] # get number of rows in extracted_rows
    object_type_match_percentage = round((num_rows-num_object_type_mismatch) / num_rows * 100,2)
    process_object_type_match_percentage = round((num_rows-num_process_object_type_mismatch) / num_rows * 100,2)
    process_object_parameter_match_percentage = round((num_rows-num_process_object_parameter_mismatch) / num_rows * 100,2)

    print(f"Number of successful mappings:")
    print(f"Signal Data Type: {num_rows-num_object_type_mismatch}/{num_rows} ({object_type_match_percentage}%)")
    print(f"Function Block Type: {num_rows-num_process_object_type_mismatch}/{num_rows} ({process_object_type_match_percentage}%)")
    print(f"Function Block Parameter: {num_rows-num_process_object_parameter_mismatch}/{num_rows} ({process_object_parameter_match_percentage}%)")
    print("Ground truth check done.")

    with open(result_filename, "w") as f:
        f.write("=== GROUND TRUTH COMPARISON RESULTS ===\n\n")
        f.write(f"Signal Data Type: {num_rows-num_object_type_mismatch}/{num_rows} ({object_type_match_percentage}%)\n")
        f.write(f"Function Block Type: {num_rows-num_process_object_type_mismatch}/{num_rows} ({process_object_type_match_percentage}%)\n")
        f.write(f"Function Block Parameter: {num_rows-num_process_object_parameter_mismatch}/{num_rows} ({process_object_parameter_match_percentage}%)\n\n")
        
        # Write all mismatch messages to the file
        if mismatch_messages:
            f.write("=== DETAILED MISMATCH INFORMATION ===\n\n")
            for message in mismatch_messages:
                f.write(f"{message}\n")
        else:
            f.write("No mismatches found.\n")

def save_output_to_excel(df, input_filename):
    """Save the processed data to an Excel file with formatted headers."""
    base_filename = input_filename.split(".")[0] + "_output"
    filename = base_filename + ".xlsx"
    
    # Check if file exists and generate a unique filename if needed
    counter = 1
    while os.path.exists(filename):
        filename = f"{base_filename}{counter}.xlsx"
        counter += 1
    
    # Create a Pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Write the dataframe to the Excel file
    df.to_excel(writer, sheet_name="IO List", index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets["IO List"]
    
    # Define the header style: bold font and light gray background
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Apply the style to each cell in the header row
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    # Identify the last three columns and apply light gray background to all their cells
    total_cols = len(df.columns)
    if total_cols >= 3:
        for col_num in range(total_cols - 2, total_cols + 1):
            for row_num in range(2, len(df) + 2):  # Start from row 2 (skip header)
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Auto-fit column widths based on content
    for col_num, column in enumerate(df.columns, 1):
        # Get maximum length of column data (including header)
        max_length = max(
            df[column].astype(str).map(len).max(),  # Max length of data
            len(str(column))  # Length of column header
        ) + 2  # Add a little extra space
        
        # Set the column width
        column_letter = worksheet.cell(row=1, column=col_num).column_letter
        worksheet.column_dimensions[column_letter].width = max_length
    
    # Save and close the workbook
    writer.close()
    
    print(f"Workbook '{filename}' created successfully with formatted headers and auto-fitted columns.")
    return filename
